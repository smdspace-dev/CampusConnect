from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db.models import Q, Count, Avg
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO

from .models import Department, Course, Staff, Cluster, Club, Student, Enrollment
from .serializers import (
    DepartmentSerializer, CourseSerializer, StaffSerializer, 
    ClusterSerializer, ClubSerializer, StudentSerializer, EnrollmentSerializer
)

# Import models from other apps for dashboard stats
from common_features.models import Event, Notification, AuditLog
from placement_management.models import PlacementRecord, Company
from teacher_interface.models import Attendance, Assignment

class IsAdminRole(IsAuthenticated):
    """Custom permission class for admin-only access"""
    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        # Check if user has staff_profile with admin role
        staff = getattr(request.user, 'staff_profile', None)
        if staff and staff.role in ('admin', 'super_admin'):
            return True
        return request.user.is_superuser

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all().order_by('name')
    serializer_class = DepartmentSerializer
    permission_classes = [IsAdminRole]
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get department statistics"""
        department = self.get_object()
        stats = {
            'total_staff': Staff.objects.filter(department=department, is_active=True).count(),
            'total_students': Student.objects.filter(department=department, is_active=True).count(),
            'total_courses': Course.objects.filter(department=department, is_active=True).count(),
            'total_clubs': Club.objects.filter(department=department, is_active=True).count(),
            'total_clusters': Cluster.objects.filter(department=department, is_active=True).count(),
        }
        return Response(stats)
    
    @action(detail=True, methods=['get'])
    def courses(self, request, pk=None):
        """Get courses for this department"""
        department = self.get_object()
        courses = Course.objects.filter(department=department, is_active=True)
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data)

class CourseViewSet(viewsets.ModelViewSet):
    serializer_class = CourseSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Course.objects.filter(is_active=True).order_by('department__name', 'semester', 'name')
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        # Filter by semester
        semester = self.request.query_params.get('semester', None)
        if semester:
            queryset = queryset.filter(semester=semester)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def by_department(self, request):
        """Get courses organized by department"""
        courses_by_dept = {}
        departments = Department.objects.filter(is_active=True)
        
        for dept in departments:
            dept_courses = Course.objects.filter(department=dept, is_active=True).order_by('semester', 'name')
            courses_by_dept[dept.name] = CourseSerializer(dept_courses, many=True).data
        
        return Response(courses_by_dept)

class StaffViewSet(viewsets.ModelViewSet):
    serializer_class = StaffSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Staff.objects.select_related('user', 'department').filter(is_active=True).order_by('user__first_name')
        
        # Filter by role
        role = self.request.query_params.get('role', None)
        if role:
            queryset = queryset.filter(role=role)
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        # Search
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(employee_id__icontains=search)
            )
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Reset staff password"""
        staff = self.get_object()
        token = staff.generate_reset_token()
        
        # In production, send email with reset link
        try:
            send_mail(
                subject='Password Reset - Campus Connect',
                message=f'Your password reset token: {token}\nPlease use this to reset your password.',
                from_email=settings.EMAIL_HOST_USER or 'no-reply@campusconnect.edu',
                recipient_list=[staff.user.email],
                fail_silently=False,
            )
        except Exception as e:
            return Response({'error': f'Failed to send email: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response({'message': 'Password reset email sent successfully'})
    
    @action(detail=True, methods=['post'])
    def toggle_access(self, request, pk=None):
        """Toggle staff access"""
        staff = self.get_object()
        new_status = staff.toggle()
        return Response({'toggle_access': new_status, 'message': f'Access {"enabled" if new_status else "disabled"}'})
    
    @action(detail=False, methods=['get'])
    def roles_summary(self, request):
        """Get summary of staff by roles"""
        roles_data = {}
        for role_code, role_name in Staff.ROLE_CHOICES:
            count = Staff.objects.filter(role=role_code, is_active=True).count()
            roles_data[role_name] = count
        
        return Response(roles_data)

class ClusterViewSet(viewsets.ModelViewSet):
    serializer_class = ClusterSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Cluster.objects.select_related('department', 'mentor').filter(is_active=True).order_by('department__name', 'name')
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """Get students in this cluster"""
        cluster = self.get_object()
        students = Student.objects.filter(cluster=cluster, is_active=True)
        serializer = StudentSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def assign_mentor(self, request, pk=None):
        """Assign mentor to cluster"""
        cluster = self.get_object()
        mentor_id = request.data.get('mentor_id')
        
        try:
            mentor = Staff.objects.get(id=mentor_id, is_active=True)
            cluster.mentor = mentor
            cluster.save()
            return Response({'message': f'Mentor {mentor.user.get_full_name()} assigned successfully'})
        except Staff.DoesNotExist:
            return Response({'error': 'Mentor not found'}, status=status.HTTP_404_NOT_FOUND)

class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Student.objects.select_related('user', 'department', 'cluster').filter(is_active=True).order_by('year', 'roll_number')
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        # Filter by year
        year = self.request.query_params.get('year', None)
        if year:
            queryset = queryset.filter(year=year)
        
        # Filter by cluster
        cluster = self.request.query_params.get('cluster', None)
        if cluster:
            queryset = queryset.filter(cluster=cluster)
        
        # Search
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(roll_number__icontains=search)
            )
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get student statistics"""
        stats = {
            'total_students': Student.objects.filter(is_active=True).count(),
            'by_year': {},
            'by_department': {},
            'by_gender': {},
        }
        
        # Students by year
        for year_choice in Student.YEAR_CHOICES:
            year_num = year_choice[0]
            stats['by_year'][year_choice[1]] = Student.objects.filter(year=year_num, is_active=True).count()
        
        # Students by department
        departments = Department.objects.filter(is_active=True)
        for dept in departments:
            stats['by_department'][dept.name] = Student.objects.filter(department=dept, is_active=True).count()
        
        # Students by gender
        for gender_choice in Student.GENDER_CHOICES:
            gender_code = gender_choice[0]
            stats['by_gender'][gender_choice[1]] = Student.objects.filter(gender=gender_code, is_active=True).count()
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def assign_cluster(self, request, pk=None):
        """Assign student to cluster"""
        student = self.get_object()
        cluster_id = request.data.get('cluster_id')
        
        try:
            cluster = Cluster.objects.get(id=cluster_id, is_active=True)
            student.cluster = cluster
            student.save()
            return Response({'message': f'Student assigned to cluster {cluster.name} successfully'})
        except Cluster.DoesNotExist:
            return Response({'error': 'Cluster not found'}, status=status.HTTP_404_NOT_FOUND)

class ClubViewSet(viewsets.ModelViewSet):
    serializer_class = ClubSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Club.objects.select_related('coordinator', 'department').filter(is_active=True).order_by('name')
        
        # Filter by club type
        club_type = self.request.query_params.get('club_type', None)
        if club_type:
            queryset = queryset.filter(club_type=club_type)
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        return queryset
    
    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """Get club members"""
        club = self.get_object()
        from student_interface.models import ClubMembership
        memberships = ClubMembership.objects.filter(club=club, is_active=True).select_related('student__user')
        
        members_data = []
        for membership in memberships:
            members_data.append({
                'id': membership.student.id,
                'roll_number': membership.student.roll_number,
                'name': membership.student.user.get_full_name(),
                'position': membership.position,
                'joined_date': membership.joined_date,
            })
        
        return Response(members_data)
    
    @action(detail=True, methods=['post'])
    def assign_coordinator(self, request, pk=None):
        """Assign coordinator to club"""
        club = self.get_object()
        coordinator_id = request.data.get('coordinator_id')
        
        try:
            coordinator = Staff.objects.get(id=coordinator_id, is_active=True)
            club.coordinator = coordinator
            club.save()
            return Response({'message': f'Coordinator {coordinator.user.get_full_name()} assigned successfully'})
        except Staff.DoesNotExist:
            return Response({'error': 'Staff member not found'}, status=status.HTTP_404_NOT_FOUND)

class EnrollmentViewSet(viewsets.ModelViewSet):
    serializer_class = EnrollmentSerializer
    permission_classes = [IsAdminRole]
    
    def get_queryset(self):
        queryset = Enrollment.objects.select_related('student__user', 'course').filter(is_active=True).order_by('-academic_year', 'semester')
        
        # Filter by academic year
        academic_year = self.request.query_params.get('academic_year', None)
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
        
        # Filter by semester
        semester = self.request.query_params.get('semester', None)
        if semester:
            queryset = queryset.filter(semester=semester)
        
        # Filter by student
        student = self.request.query_params.get('student', None)
        if student:
            queryset = queryset.filter(student=student)
        
        # Filter by course
        course = self.request.query_params.get('course', None)
        if course:
            queryset = queryset.filter(course=course)
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def bulk_enroll(self, request):
        """Bulk enroll students in courses"""
        student_ids = request.data.get('student_ids', [])
        course_ids = request.data.get('course_ids', [])
        semester = request.data.get('semester')
        academic_year = request.data.get('academic_year')
        
        if not all([student_ids, course_ids, semester, academic_year]):
            return Response({'error': 'All fields required'}, status=status.HTTP_400_BAD_REQUEST)
        
        enrollments_created = 0
        errors = []
        
        for student_id in student_ids:
            try:
                student = Student.objects.get(id=student_id, is_active=True)
                for course_id in course_ids:
                    try:
                        course = Course.objects.get(id=course_id, is_active=True)
                        enrollment, created = Enrollment.objects.get_or_create(
                            student=student,
                            course=course,
                            semester=semester,
                            academic_year=academic_year
                        )
                        if created:
                            enrollments_created += 1
                    except Course.DoesNotExist:
                        errors.append(f'Course {course_id} not found')
            except Student.DoesNotExist:
                errors.append(f'Student {student_id} not found')
        
        return Response({
            'enrollments_created': enrollments_created,
            'errors': errors
        })

class StudentBulkUploadView(APIView):
    """Bulk upload students from Excel file"""
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAdminRole]

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Excel file required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(filename=file_obj, read_only=True)
            sheet = wb.active
            
            created_students = []
            errors = []
            
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx == 1:  # Skip header row
                    continue
                
                try:
                    # Expected columns: roll_number, first_name, last_name, email, phone, department_code, year, semester
                    roll_number = row[0]
                    first_name = row[1] or ''
                    last_name = row[2] or ''
                    email = row[3] or f'{roll_number}@college.edu'
                    phone = row[4] or ''
                    department_code = row[5]
                    year = int(row[6]) if row[6] else 1
                    semester = int(row[7]) if row[7] else 1
                    
                    if not roll_number:
                        errors.append(f'Row {idx}: Roll number required')
                        continue
                    
                    # Check if student already exists
                    if Student.objects.filter(roll_number=roll_number).exists():
                        errors.append(f'Row {idx}: Student {roll_number} already exists')
                        continue
                    
                    # Get department
                    try:
                        department = Department.objects.get(code=department_code, is_active=True)
                    except Department.DoesNotExist:
                        errors.append(f'Row {idx}: Department {department_code} not found')
                        continue
                    
                    # Create user
                    username = roll_number.lower()
                    if User.objects.filter(username=username).exists():
                        username = f"{username}_{idx}"
                    
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password='student123',  # Default password
                        first_name=first_name,
                        last_name=last_name
                    )
                    
                    # Create student
                    student = Student.objects.create(
                        user=user,
                        roll_number=roll_number,
                        department=department,
                        year=year,
                        semester=semester,
                        phone=phone
                    )
                    
                    created_students.append({
                        'roll_number': roll_number,
                        'name': f'{first_name} {last_name}',
                        'email': email
                    })
                    
                    # Send welcome email
                    try:
                        send_mail(
                            subject='Welcome to Campus Connect',
                            message=f'Dear {first_name},\n\nYour account has been created:\nUsername: {username}\nPassword: student123\n\nPlease login and change your password.',
                            from_email=settings.EMAIL_HOST_USER or 'no-reply@campusconnect.edu',
                            recipient_list=[email],
                            fail_silently=True,
                        )
                    except Exception:
                        pass  # Email sending is optional
                    
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
            
            return Response({
                'message': f'Successfully created {len(created_students)} students',
                'created_students': created_students,
                'errors': errors
            })
            
        except Exception as e:
            return Response({'error': f'Failed to process file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        dept = self.get_object()
        dept.is_active = not dept.is_active
        dept.save()
        return Response({'is_active': dept.is_active})

class StaffViewSet(viewsets.ModelViewSet):
    queryset = Staff.objects.select_related('user','department').all()
    serializer_class = StaffSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        staff = self.get_object()
        token = staff.generate_reset_token()
        # In production, send email with reset link
        try:
            send_mail(
                subject='Password Reset',
                message=f'Use this token to reset password: {token}',
                from_email=settings.EMAIL_HOST_USER or 'no-reply@example.com',
                recipient_list=[staff.user.email],
                fail_silently=True,
            )
        except Exception:
            pass
        return Response({'token': token})

    @action(detail=True, methods=['post'])
    def toggle_access(self, request, pk=None):
        staff = self.get_object()
        new_status = staff.toggle()
        return Response({'toggle_access': new_status})

class ClubViewSet(viewsets.ModelViewSet):
    queryset = Club.objects.prefetch_related('members').all()
    serializer_class = ClubSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def toggle_member(self, request, pk=None):
        club = self.get_object()
        student_id = request.data.get('student_id')
        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            return Response({'detail': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
        if student in club.members.all():
            club.members.remove(student)
            return Response({'status': 'removed'})
        else:
            club.members.add(student)
            return Response({'status': 'added'})

class StudentBulkUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'File required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(filename=file_obj, read_only=True)
            sheet = wb.active
            created = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx == 0:
                    continue
                username = row[0] or f'student{idx}'
                email = row[1] or f'{username}@example.com'
                first_name = row[2] or ''
                last_name = row[3] or ''
                if User.objects.filter(username=username).exists():
                    continue
                user = User.objects.create_user(username=username, email=email, password='password123', first_name=first_name, last_name=last_name)
                student = Student.objects.create(user=user, bulk_created=True)
                created.append(user.username)
                # Optionally send email
                try:
                    send_mail(
                        subject='Welcome to College System',
                        message=f'Your account {username} has been created. Default password: password123',
                        from_email=settings.EMAIL_HOST_USER or 'no-reply@example.com',
                        recipient_list=[email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
            return Response({'created': created})
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# Dashboard Statistics API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard_stats(request):
    """Get comprehensive dashboard statistics for admin"""
    try:
        # Basic counts
        total_students = Student.objects.filter(is_active=True).count()
        total_faculty = Staff.objects.filter(is_active=True).count()
        total_departments = Department.objects.filter(is_active=True).count()
        total_courses = Course.objects.filter(is_active=True).count()
        total_clubs = Club.objects.filter(is_active=True).count()
        
        # Additional stats
        total_placements = PlacementRecord.objects.filter(is_confirmed=True).count()
        total_companies = Company.objects.filter(is_active=True).count()
        total_events = Event.objects.count()
        
        # Recent activities (last 7 days)
        week_ago = timezone.now() - timedelta(days=7)
        recent_enrollments = Enrollment.objects.filter(enrolled_at__gte=week_ago).count()
        recent_logins = AuditLog.objects.filter(action='login', timestamp__gte=week_ago).count()
        
        # Student statistics by year
        students_by_year = {}
        for year_choice in Student.YEAR_CHOICES:
            year_num = year_choice[0]
            count = Student.objects.filter(year=year_num, is_active=True).count()
            students_by_year[year_choice[1]] = count
        
        # Department-wise student count
        department_stats = []
        for dept in Department.objects.filter(is_active=True):
            student_count = Student.objects.filter(department=dept, is_active=True).count()
            staff_count = Staff.objects.filter(department=dept, is_active=True).count()
            department_stats.append({
                'name': dept.name,
                'code': dept.code,
                'students': student_count,
                'staff': staff_count
            })
        
        # Recent activities for timeline
        recent_activities = []
        
        # Recent student registrations
        recent_students = Student.objects.filter(
            created_at__gte=week_ago
        ).select_related('user').order_by('-created_at')[:5]
        
        for student in recent_students:
            recent_activities.append({
                'type': 'student_registration',
                'title': 'New Student Registration',
                'description': f'{student.user.get_full_name()} ({student.roll_number}) joined',
                'timestamp': student.created_at,
                'icon': 'fa-user-plus'
            })
        
        # Recent events
        recent_events = Event.objects.filter(
            created_at__gte=week_ago
        ).order_by('-created_at')[:3]
        
        for event in recent_events:
            recent_activities.append({
                'type': 'event_created',
                'title': 'New Event Created',
                'description': f'{event.title} scheduled for {event.start_date.date()}',
                'timestamp': event.created_at,
                'icon': 'fa-calendar-plus'
            })
        
        # Sort activities by timestamp
        recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Quick stats for cards
        stats = {
            'overview': {
                'total_students': total_students,
                'total_faculty': total_faculty,
                'total_departments': total_departments,
                'total_courses': total_courses,
                'total_clubs': total_clubs,
                'total_placements': total_placements,
                'total_companies': total_companies,
                'total_events': total_events
            },
            'recent_metrics': {
                'new_enrollments_week': recent_enrollments,
                'active_sessions_week': recent_logins,
                'pending_applications': 0,  # Can be calculated from placement apps
                'upcoming_events': Event.objects.filter(
                    start_date__gte=timezone.now(),
                    start_date__lte=timezone.now() + timedelta(days=7)
                ).count()
            },
            'students_by_year': students_by_year,
            'department_stats': department_stats,
            'recent_activities': recent_activities[:10],  # Limit to 10 activities
            'system_health': {
                'database_status': 'healthy',
                'active_users_today': AuditLog.objects.filter(
                    action='login',
                    timestamp__date=timezone.now().date()
                ).values('user').distinct().count(),
                'server_uptime': 'Good',
                'last_backup': 'Not configured'  # Can be implemented later
            }
        }
        
        return Response(stats)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to generate dashboard stats: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Academic Management APIs
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def academic_overview(request):
    """Get academic overview statistics"""
    try:
        current_year = timezone.now().year
        academic_year = f"{current_year}-{current_year + 1}"
        
        # Course statistics
        course_stats = {
            'total_courses': Course.objects.filter(is_active=True).count(),
            'courses_by_type': {},
            'courses_by_semester': {},
            'enrollment_stats': {}
        }
        
        # Courses by type
        for course_type_choice in Course.COURSE_TYPE_CHOICES:
            type_code = course_type_choice[0]
            count = Course.objects.filter(course_type=type_code, is_active=True).count()
            course_stats['courses_by_type'][course_type_choice[1]] = count
        
        # Courses by semester
        for semester in range(1, 9):
            count = Course.objects.filter(semester=semester, is_active=True).count()
            course_stats['courses_by_semester'][f'Semester {semester}'] = count
        
        # Enrollment statistics
        total_enrollments = Enrollment.objects.filter(is_active=True).count()
        current_enrollments = Enrollment.objects.filter(
            academic_year=academic_year,
            is_active=True
        ).count()
        
        course_stats['enrollment_stats'] = {
            'total_enrollments': total_enrollments,
            'current_year_enrollments': current_enrollments,
            'average_students_per_course': round(
                total_enrollments / max(course_stats['total_courses'], 1), 2
            )
        }
        
        return Response(course_stats)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to generate academic overview: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
